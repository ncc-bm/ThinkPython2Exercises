import openpyxl

wb = openpyxl.Workbook() # create a blank workbook
print(wb.sheetnames)
print()

sheet = wb.active
print(sheet.title)
print()

sheet.title = 'Spam Bacon Eggs Sheet' # Change title
print(wb.sheetnames)
print()

""" 
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')
 """

wb.create_sheet()
print(wb.sheetnames)
print()

wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
print()

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
print()

del wb['Middle Sheet']
print(wb.sheetnames)
print()

del wb['Sheet']
print(wb.sheetnames)
print()

print(wb.active)
sheet = wb.active
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)
wb.save('example_copy.xlsx')
