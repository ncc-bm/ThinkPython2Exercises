import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')

""" 
print(type(wb))

print(wb.sheetnames)

sheet = wb['Sheet3']
print(sheet)
print(type(sheet))
print(sheet.title)

anotherSheet = wb.active
print(anotherSheet)
print(anotherSheet.title)
 """

sheet = wb['Sheet1']
print(sheet['A1'])
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Row %s, Column %s is %s' % (c.row, get_column_letter(c.column), c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print(sheet['C1'].value)

print()
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print()
print(sheet.max_row)
print(sheet.max_column)

print()
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(c.column))

print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))


print(tuple(sheet['A1': 'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END of ROW ---')
    