import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) > 1:
    tableSize = sys.argv[1]
else:
    tableSize = input('Please enter table size: ')

print('You entered %s as the table size' % tableSize)

wb = openpyxl.Workbook()
sheet = wb.active

fontObj = Font(bold = True)

for i in range(1, int(tableSize)+1):
    sheet.cell(row=1, column=i+1).font = fontObj
    sheet.cell(row=1, column=i+1).value = i

for i in range(1, int(tableSize)+1):
    sheet.cell(row=i+1, column=1).font = fontObj
    sheet.cell(row=i+1, column=1).value = i

for i in range(2, int(tableSize)+2):
    for j in range(2, int(tableSize)+2):
        sheet.cell(row=i, column=j).value = (i-1)*(j-1)

wb.save('multiplication_table.xlsx')
