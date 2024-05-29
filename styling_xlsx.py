import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()

sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello world!'
#wb.save('styles.xlsx')

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['B1'].font = fontObj1
sheet['B1'] = 'Bold Time New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

sheet['A4'] = 200
sheet['A5'] = 300
sheet['A6'] = '=SUM(A4:A5)' # Set the formula

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

# Set the hight and width:
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['C'].width = 20

# merging cells
sheet.merge_cells('A6:D9') # Merge all these cells.
sheet['A6'] = 'Twelve cells merged together.'
sheet.merge_cells('C10:D10') # Merge these two cells.
sheet['C10'] = 'Two merged cells.'

sheet.unmerge_cells('A6:D9') # Split these cells up.
sheet.unmerge_cells('C10:d10')
wb.save('styles.xlsx')

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2' # Freeze the rows above A2.
wb.save('freezeExample.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
for i in range (1, 11): # create some data in column A
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')